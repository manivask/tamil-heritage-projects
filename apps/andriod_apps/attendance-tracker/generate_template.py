import os
import sys

def create_excel():
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl

    wb = openpyxl.Workbook()
    
    locations = ["Chennai Cholai", "Madurai Malar", "Kovai Kani", "Nellai Neer"]
    grades = ["LKG", "UKG", "Grade 1", "Grade 2", "Grade 3", "Grade 4", "Grade 5", "Grade 6", "Grade 7", "Grade 8"]
    
    first_names = ["Aadhavan", "Abinaya", "Akilan", "Amudhan", "Anbarasu", "Arul", "Balan", "Chitra", "Devan", "Ezhil", 
                   "Iniyan", "Kailash", "Kamali", "Kavin", "Kumaran", "Madhi", "Mugilan", "Nila", "Oviya", "Pugazh", 
                   "Saran", "Tamilarasan", "Thenmozhi", "Valan", "Yazhini", "Venkatesh", "Sundar", "Rajesh", "Karthik", "Divya"]
    last_names = ["Kumar", "Selvam", "Rajan", "Sundaram", "Pandian", "Arasan", "Thambi", "Vasagam", "Nambi", "Govindan", "Mani", "Raja"]

    # 1. Teachers Sheet
    ws_teachers = wb.active
    ws_teachers.title = "Teachers"
    ws_teachers.append(["ID", "Name", "Location", "Class Assignment", "Email", "Phone"])
    
    t_count = 1
    for loc in locations:
        for gr in grades:
            for t_num in range(1, 3):
                fname = first_names[(t_count * 3) % len(first_names)]
                lname = last_names[(t_count * 7) % len(last_names)]
                name = f"{fname} {lname}"
                email = f"{fname.lower()}.{lname.lower()}@school.com"
                phone = f"987-654-3{t_count:03d}"
                ws_teachers.append([f"T{t_count:03d}", name, loc, gr, email, phone])
                t_count += 1
                
    # 2. Students Sheet
    ws_students = wb.create_sheet(title="Students")
    ws_students.append(["ID", "Name", "Location", "Grade", "Parent Email"])
    
    # Generate 90 kids distributed across 4 locations and 10 grades
    # Clean First Name + Last Name (no initials)
    s_count = 1
    for i in range(90):
        fname = first_names[(i * 4) % len(first_names)]
        lname = last_names[(i * 5) % len(last_names)]
        name = f"{fname} {lname}"
        loc = locations[i % len(locations)]
        grade = grades[i % len(grades)]
        ws_students.append([f"S{s_count:03d}", name, loc, grade, f"parent{s_count}@example.com"])
        s_count += 1
        
    # 3. Committee Sheet (Registry)
    ws_committee = wb.create_sheet(title="Committee")
    ws_committee.append(["ID", "Name", "Role", "Authorized PIN"])
    
    committee_members = [
        ["C01", "Senthamil Arasan", "President (Committee 1)", "9900"],
        ["C02", "Kavin Selvam", "Principal (Chennai Cholai)", "9001"],
        ["C03", "Ezhil Tamilarasan", "Principal (Madurai Malar)", "9002"],
        ["C04", "Kailash Balan", "Principal (Kovai Kani)", "9003"],
        ["C05", "Mugilan Pugazh", "Principal (Nellai Neer)", "9004"],
        ["C06", "Amudha Kumar", "Vice Principal (Chennai Cholai)", "9101"],
        ["C07", "Kamali Chitra", "Vice Principal (Madurai Malar)", "9102"],
        ["C08", "Yazhini Nila", "Vice Principal (Kovai Kani)", "9103"],
        ["C09", "Oviya Thenmozhi", "Vice Principal (Nellai Neer)", "9104"],
        ["C10", "Bharathi Raja", "Committee Member 2", "1001"],
        ["C11", "Elango Mani", "Committee Member 3", "1002"],
        ["C12", "Kavitha Sundar", "Committee Member 4", "1003"],
        ["C13", "Muthu Pandian", "Committee Member 5", "1004"],
        ["C14", "Nila Govindan", "Committee Member 6", "1005"],
        ["C15", "Selvam Nambi", "Committee Member 7", "1006"],
        ["C16", "Senthamil Thambi", "Committee Member 8", "1007"]
    ]
    for row in committee_members:
        ws_committee.append(row)

    output_path = "attendance_template.xlsx"
    wb.save(output_path)
    print(f"Excel template generated successfully at: {os.path.abspath(output_path)}")

if __name__ == "__main__":
    create_excel()
